import openpyxl

# Load MFD
wb_mfd = openpyxl.load_workbook(r"c:\laragon\www\dashboard-se2026\data\mfd\sm22025\mfd_25_2_3509.xlsx", data_only=True)
ws_mfd = wb_mfd['Sheet 1']

mfd_desas = set()
mfd_kecamatans = set()
mfd_desa_by_kec = {}

for row in ws_mfd.iter_rows(min_row=2, values_only=True):
    kdkec = str(row[7]) # kdkec
    kddesa = str(row[9]) # kddesa
    nmkec = str(row[8]) # nmkec
    nmdesa = str(row[10]) # nmdesa
    
    mfd_desas.add((kdkec, kddesa))
    mfd_kecamatans.add(kdkec)
    
    if kdkec not in mfd_desa_by_kec:
        mfd_desa_by_kec[kdkec] = set()
    mfd_desa_by_kec[kdkec].add(kddesa)

wb_mfd.close()

# Load MSUBSLS
wb_sub = openpyxl.load_workbook(r"c:\laragon\www\dashboard-se2026\data\mfd\sm22025\msubsls_25_2_3509.xlsx", data_only=True)
ws_sub = wb_sub['Sheet 1']

sub_desas = set()
sub_kecamatans = set()
sub_desa_by_kec = {}

for row in ws_sub.iter_rows(min_row=2, values_only=True):
    kdkec = str(row[10]) # kdkec
    kddesa = str(row[12]) # kddesa
    
    sub_desas.add((kdkec, kddesa))
    sub_kecamatans.add(kdkec)
    
    if kdkec not in sub_desa_by_kec:
        sub_desa_by_kec[kdkec] = set()
    sub_desa_by_kec[kdkec].add(kddesa)

wb_sub.close()

print("=" * 80)
print("CROSS VALIDATION: MFD VS MSUBSLS")
print("=" * 80)
print(f"MFD: {len(mfd_kecamatans)} kecamatans, {len(mfd_desas)} desas")
print(f"MSUBSLS: {len(sub_kecamatans)} kecamatans, {len(sub_desas)} desas")

# Check if kecamatans match
kec_mfd_not_sub = mfd_kecamatans - sub_kecamatans
kec_sub_not_mfd = sub_kecamatans - mfd_kecamatans
print(f"\nKecamatan in MFD but not in MSUBSLS: {kec_mfd_not_sub}")
print(f"Kecamatan in MSUBSLS but not in MFD: {kec_sub_not_mfd}")

# Check if desas match
desa_mfd_not_sub = mfd_desas - sub_desas
desa_sub_not_mfd = sub_desas - mfd_desas
print(f"\nDesa in MFD but not in MSUBSLS: {len(desa_mfd_not_sub)}")
if desa_mfd_not_sub:
    print(f"  Sample: {list(desa_mfd_not_sub)[:5]}")
    
print(f"Desa in MSUBSLS but not in MFD: {len(desa_sub_not_mfd)}")
if desa_sub_not_mfd:
    print(f"  Sample: {list(desa_sub_not_mfd)[:5]}")

print("\nValidation completed!")
