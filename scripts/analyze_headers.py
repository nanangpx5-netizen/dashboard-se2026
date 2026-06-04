import openpyxl
import json

# Analyze MFD file headers and first few rows
wb = openpyxl.load_workbook(r"c:\laragon\www\dashboard-se2026\data\mfd\sm22025\mfd_25_2_3509.xlsx", data_only=True)

print("=" * 80)
print("MFD FILE - Sheet names:", wb.sheetnames)
print("=" * 80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    print(f"Dimensions: {ws.dimensions}, Max Row: {ws.max_row}, Max Col: {ws.max_column}")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True), start=1):
        print(f"Row {row_idx}: {list(row)}")

wb.close()

# Analyze MSUBSLS file headers  
print("\n" + "=" * 80)
print("MSUBSLS FILE")
print("=" * 80)

wb2 = openpyxl.load_workbook(r"c:\laragon\www\dashboard-se2026\data\mfd\sm22025\msubsls_25_2_3509.xlsx", data_only=True)
print("Sheet names:", wb2.sheetnames)

for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    print(f"Dimensions: {ws.dimensions}, Max Row: {ws.max_row}, Max Col: {ws.max_column}")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True), start=1):
        print(f"Row {row_idx}: {list(row)}")

wb2.close()
