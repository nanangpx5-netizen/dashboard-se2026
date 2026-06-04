import openpyxl
import sys
import os

def analyze_xlsx(file_path):
    if not os.path.exists(file_path):
        print(f"Error: File {file_path} not found.")
        return

    print("=" * 80)
    print(f"FILE: {os.path.basename(file_path)}")
    print("=" * 80)

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        print("Sheet names:", wb.sheetnames)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")
            
            # Print first 5 rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
                print(f"Row {row_idx}: {list(row)}")
        
        wb.close()
    except Exception as e:
        print(f"Error processing file: {e}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python analyze_xlsx_generic.py <file_path>")
    else:
        analyze_xlsx(sys.argv[1])
