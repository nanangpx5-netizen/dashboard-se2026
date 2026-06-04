import openpyxl
import json
from collections import Counter

# ============================================================
# MFD FILE - Deep Analysis
# ============================================================
wb = openpyxl.load_workbook(r"c:\laragon\www\dashboard-se2026\data\mfd\sm22025\mfd_25_2_3509.xlsx", data_only=True)
ws = wb['Sheet 1']

print("=" * 80)
print("MFD FILE - DEEP ANALYSIS")
print("=" * 80)
print(f"Total rows (excl header): {ws.max_row - 1}")

# Parse all data
mfd_data = []
headers = [cell.value for cell in ws[1]]
print(f"Headers: {headers}")

kecamatan_set = set()
desa_count = 0
klas_counter = Counter()
status_counter = Counter()
pem_counter = Counter()

total_subsls = 0
total_sls = 0
total_nonsls = 0
total_bs = 0

kec_stats = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    desa_count += 1
    kecamatan = row[8]  # nmkec
    kdkec = row[7]  # kdkec
    desa = row[10]  # nmdesa
    klas = row[11]  # klas
    status = row[12]  # status_desa
    pem = row[13]  # status_pem
    jml_subsls = row[16] or 0
    jml_sls = row[17] or 0
    jml_nonsls = row[18] or 0
    jml_bs = row[19] or 0
    lat = row[14]
    lon = row[15]
    
    kecamatan_set.add((kdkec, kecamatan))
    klas_counter[klas] += 1
    status_counter[status] += 1
    pem_counter[pem] += 1
    
    total_subsls += jml_subsls
    total_sls += jml_sls
    total_nonsls += jml_nonsls
    total_bs += jml_bs
    
    key = f"{kdkec}-{kecamatan}"
    if key not in kec_stats:
        kec_stats[key] = {"kd": kdkec, "nm": kecamatan, "desa": 0, "subsls": 0, "sls": 0, "nonsls": 0, "bs": 0, "perkotaan": 0, "pedesaan": 0}
    kec_stats[key]["desa"] += 1
    kec_stats[key]["subsls"] += jml_subsls
    kec_stats[key]["sls"] += jml_sls
    kec_stats[key]["nonsls"] += jml_nonsls
    kec_stats[key]["bs"] += jml_bs
    if klas == 1:
        kec_stats[key]["perkotaan"] += 1
    else:
        kec_stats[key]["pedesaan"] += 1

wb.close()

print(f"\nTotal Desa/Kelurahan: {desa_count}")
print(f"Total Kecamatan: {len(kecamatan_set)}")
print(f"\nKlasifikasi: {dict(klas_counter)} (1=Perkotaan, 2=Pedesaan)")
print(f"Status Desa: {dict(status_counter)}")
print(f"Status Pemerintahan: {dict(pem_counter)}")
print(f"\nTotal Sub-SLS: {total_subsls}")
print(f"Total SLS: {total_sls}")
print(f"Total Non-SLS: {total_nonsls}")
print(f"Total BS: {total_bs}")

print(f"\n--- Kecamatan Summary (sorted by code) ---")
for key in sorted(kec_stats.keys()):
    s = kec_stats[key]
    print(f"  {s['kd']} {s['nm']}: {s['desa']} desa, SubSLS={s['subsls']}, SLS={s['sls']}, NonSLS={s['nonsls']}, BS={s['bs']}, Kota={s['perkotaan']}, Desa={s['pedesaan']}")

# ============================================================
# MSUBSLS FILE - Deep Analysis
# ============================================================
print("\n" + "=" * 80)
print("MSUBSLS FILE - DEEP ANALYSIS")
print("=" * 80)

wb2 = openpyxl.load_workbook(r"c:\laragon\www\dashboard-se2026\data\mfd\sm22025\msubsls_25_2_3509.xlsx", data_only=True)
ws2 = wb2['Sheet 1']
print(f"Total rows (excl header): {ws2.max_row - 1}")

headers2 = [cell.value for cell in ws2[1]]
print(f"Headers: {headers2}")

jenis_counter = Counter()
klas_counter2 = Counter()
dominan_counter = Counter()
total_kk = 0
total_bstt = 0
total_bsbtt = 0
total_bsttk = 0
total_bku = 0
total_usaha = 0
total_muatan = 0

for row in ws2.iter_rows(min_row=2, values_only=True):
    jenis = row[5]  # jenis
    klas2 = row[16]  # klas
    dominan = row[24]  # dominan
    
    jenis_counter[jenis] += 1
    klas_counter2[klas2] += 1
    dominan_counter[dominan] += 1
    
    total_kk += (row[17] or 0)
    total_bstt += (row[18] or 0)
    total_bsbtt += (row[19] or 0)
    total_bsttk += (row[20] or 0)
    total_bku += (row[21] or 0)
    total_usaha += (row[22] or 0)
    total_muatan += (row[23] or 0)

wb2.close()

print(f"\nJenis SLS: {dict(jenis_counter)}")
print(f"Klasifikasi: {dict(klas_counter2)}")
print(f"Dominan: {dict(dominan_counter)}")
print(f"\nTotal KK (Kepala Keluarga): {total_kk:,}")
print(f"Total BSTT (Bangunan Sensus TT): {total_bstt:,}")
print(f"Total BSBTT (Bangunan Sensus BTT): {total_bsbtt:,}")
print(f"Total BSTTK: {total_bsttk:,}")
print(f"Total BKU (Bangunan Khusus Usaha): {total_bku:,}")
print(f"Total Usaha: {total_usaha:,}")
print(f"Total Muatan: {total_muatan:,}")

# Read keterangan sheet for MSUBSLS
ws_ket = wb2 if False else openpyxl.load_workbook(r"c:\laragon\www\dashboard-se2026\data\mfd\sm22025\msubsls_25_2_3509.xlsx", data_only=True)['keterangan']
print(f"\n--- Keterangan Sheet ---")
for row in ws_ket.iter_rows(min_row=1, values_only=True):
    print(f"  {list(row)}")
