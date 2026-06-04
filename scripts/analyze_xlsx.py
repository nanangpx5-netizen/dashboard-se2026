import openpyxl
import json
import sys

def analyze_xlsx(filepath, max_rows=50):
    """Analyze an Excel file and output its structure and sample data."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {
        "file": filepath,
        "sheets": []
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "headers": [],
            "sample_data": [],
            "merged_cells": [str(m) for m in ws.merged_cells.ranges]
        }
        
        # Get all rows up to max_rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=False), start=1):
            row_data = []
            for cell in row:
                val = cell.value
                if val is not None:
                    val = str(val)
                row_data.append(val)
            
            if row_idx <= 5:
                sheet_info["headers"].append({"row": row_idx, "values": row_data})
            else:
                sheet_info["sample_data"].append({"row": row_idx, "values": row_data})
        
        result["sheets"].append(sheet_info)
    
    return result

# Analyze MFD file
print("=" * 80)
print("ANALYZING MFD FILE (Master File Desa)")
print("=" * 80)
mfd = analyze_xlsx(r"c:\laragon\www\dashboard-se2026\data\mfd\sm22025\mfd_25_2_3509.xlsx", max_rows=30)
print(json.dumps(mfd, indent=2, ensure_ascii=False, default=str))

print("\n" + "=" * 80)
print("ANALYZING MSUBSLS FILE (Master Sub-SLS)")
print("=" * 80)
msub = analyze_xlsx(r"c:\laragon\www\dashboard-se2026\data\mfd\sm22025\msubsls_25_2_3509.xlsx", max_rows=30)
print(json.dumps(msub, indent=2, ensure_ascii=False, default=str))
